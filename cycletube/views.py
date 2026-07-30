import datetime
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.shortcuts import render, redirect

from .models import (
    CycleTubeItem, CycleTubeEntry, CycleTubeDailyManualEntry,
    BUCKET_CHOICES, PACK_FACTOR, VB_FACTOR, COMB_FACTOR,
)
from .forms import (
    CycleTubeItemForm, ProductionEntryForm, SaleEntryForm,
    AdjustmentEntryForm, CycleTubeDailyManualEntryForm,
)


@login_required
def dashboard(request):
    q = request.GET.get("q", "").strip()
    items = CycleTubeItem.objects.filter(is_active=True)
    if q:
        items = items.filter(
            Q(size__icontains=q) | Q(type__icontains=q) | Q(brand__icontains=q)
        )

    today = datetime.date.today()
    today_production = CycleTubeEntry.objects.filter(entry_type="production", date=today).aggregate(t=Sum("quantity"))["t"] or 0
    today_sale = CycleTubeEntry.objects.filter(entry_type="sale", date=today).aggregate(t=Sum("quantity"))["t"] or 0

    month_start = today.replace(day=1)
    month_production = CycleTubeEntry.objects.filter(entry_type="production", date__gte=month_start, date__lte=today).aggregate(t=Sum("quantity"))["t"] or 0
    month_sale = CycleTubeEntry.objects.filter(entry_type="sale", date__gte=month_start, date__lte=today).aggregate(t=Sum("quantity"))["t"] or 0

    for item in items:
        item.month_production = CycleTubeEntry.objects.filter(
            tube_item=item, entry_type="production", date__gte=month_start, date__lte=today
        ).aggregate(t=Sum("quantity"))["t"] or 0
        item.month_sale = CycleTubeEntry.objects.filter(
            tube_item=item, entry_type="sale", date__gte=month_start, date__lte=today
        ).aggregate(t=Sum("quantity"))["t"] or 0

    grand_total = sum(item.total_stock for item in items)
    total_stock_col = sum(item.stock for item in items)
    total_rfm_col = sum(item.rfm_stock for item in items)

    context = {
        "items": items,
        "q": q,
        "today": today,
        "today_production": today_production,
        "today_sale": today_sale,
        "month_production": month_production,
        "month_sale": month_sale,
        "grand_total": grand_total,
        "total_stock_col": total_stock_col,
        "total_rfm_col": total_rfm_col,
    }
    return render(request, "cycletube/dashboard.html", context)


@login_required
def add_item(request):
    form = CycleTubeItemForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.size = item.size.strip()
        item.type = item.type.strip()
        item.brand = item.brand.strip()
        item.save()
        messages.success(request, f"Cycle Tube '{item}' add ho gaya.")
        return redirect("tube_dashboard")
    return render(request, "cycletube/add_item.html", {"form": form})


@login_required
def add_production(request):
    form = ProductionEntryForm(request.POST or None, initial={"date": datetime.date.today()})
    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            item = form.cleaned_data["tube_item"]
            qty = form.cleaned_data["quantity"]
            tube_quality = form.cleaned_data.get("tube_quality", "normal")
            item.stock += qty
            item.save(update_fields=["stock"])
            CycleTubeEntry.objects.create(
                tube_item=item,
                entry_type="production",
                bucket="stock",
                quantity=qty,
                tube_quality=tube_quality,
                date=form.cleaned_data["date"],
                remark=form.cleaned_data["remark"],
                user=request.user,
            )
        messages.success(request, f"Production entry added: {item} +{qty} ({tube_quality})")
        return redirect("tube_add_production")
    recent = CycleTubeEntry.objects.filter(entry_type="production").select_related("tube_item", "user")[:15]
    return render(request, "cycletube/add_production.html", {"form": form, "recent": recent})


@login_required
def add_sale(request):
    form = SaleEntryForm(request.POST or None, initial={"date": datetime.date.today()})
    if request.method == "POST" and form.is_valid():
        item = form.cleaned_data["tube_item"]
        bucket = form.cleaned_data["bucket"]
        qty = form.cleaned_data["quantity"]
        bill_number = form.cleaned_data["bill_number"].strip()
        current = getattr(item, bucket)

        duplicate = CycleTubeEntry.objects.filter(
            entry_type="sale", bill_number__iexact=bill_number
        ).first()

        if duplicate:
            messages.error(
                request,
                f"Bill number '{bill_number}' Already exists in sale entry: "
                f"({duplicate.date} - {duplicate.tube_item}). Duplicate Entry not allowed, "
                f"bill number check kar lo."
            )
        elif qty > current:
            messages.error(request, f"'{item}' ke '{dict(BUCKET_CHOICES)[bucket]}' bucket mai sirf {current} hi available hai.")
        else:
            with transaction.atomic():
                setattr(item, bucket, current - qty)
                item.save(update_fields=[bucket])
                CycleTubeEntry.objects.create(
                    tube_item=item,
                    entry_type="sale",
                    bucket=bucket,
                    quantity=qty,
                    date=form.cleaned_data["date"],
                    bill_number=bill_number,
                    remark=form.cleaned_data["remark"],
                    user=request.user,
                )
            messages.success(request, f"Sale entry added: {item} -{qty} ({dict(BUCKET_CHOICES)[bucket]}) | Bill: {bill_number}")
            return redirect("tube_add_sale")
    recent = CycleTubeEntry.objects.filter(entry_type="sale").select_related("tube_item", "user")[:15]
    return render(request, "cycletube/add_sale.html", {"form": form, "recent": recent})


@login_required
def add_adjustment(request):
    form = AdjustmentEntryForm(request.POST or None, initial={"date": datetime.date.today()})
    if request.method == "POST" and form.is_valid():
        item = form.cleaned_data["tube_item"]
        bucket = form.cleaned_data["bucket"]
        action = form.cleaned_data["action"]
        qty = form.cleaned_data["quantity"]
        current = getattr(item, bucket)
        if action == "subtract" and qty > current:
            messages.error(request, f"'{item}' ke '{dict(BUCKET_CHOICES)[bucket]}' Only {current} Items are available in Bucket.")
        else:
            signed_qty = qty if action == "add" else -qty
            with transaction.atomic():
                setattr(item, bucket, current + signed_qty)
                item.save(update_fields=[bucket])
                CycleTubeEntry.objects.create(
                    tube_item=item,
                    entry_type="adjustment",
                    bucket=bucket,
                    quantity=signed_qty,
                    date=form.cleaned_data["date"],
                    remark=form.cleaned_data["remark"],
                    user=request.user,
                )
            messages.success(request, f"Adjustment added: {item} {signed_qty:+d} ({dict(BUCKET_CHOICES)[bucket]})")
            return redirect("tube_add_adjustment")
    return render(request, "cycletube/add_adjustment.html", {"form": form})


@login_required
def entries_log(request):
    date_str = request.GET.get("date", "")
    month_str = request.GET.get("month", "")
    entry_type = request.GET.get("type", "")

    entries = CycleTubeEntry.objects.select_related("tube_item", "user")

    if date_str:
        try:
            d = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            entries = entries.filter(date=d)
        except ValueError:
            pass
    elif month_str:
        try:
            y, m = month_str.split("-")
            entries = entries.filter(date__year=int(y), date__month=int(m))
        except ValueError:
            pass

    if entry_type:
        entries = entries.filter(entry_type=entry_type)

    entries = entries[:500]

    context = {
        "entries": entries,
        "date_str": date_str,
        "month_str": month_str,
        "entry_type": entry_type,
    }
    return render(request, "cycletube/entries_log.html", context)


@login_required
def monthly_report(request):
    month_str = request.GET.get("month", "")
    today = datetime.date.today()
    if month_str:
        try:
            y, m = month_str.split("-")
            year, month = int(y), int(m)
        except ValueError:
            year, month = today.year, today.month
    else:
        year, month = today.year, today.month

    items = CycleTubeItem.objects.filter(is_active=True)
    rows = []

    total_production = 0
    total_sale = 0
    total_stock = 0
    grand_total = 0

    for item in items:
        production = CycleTubeEntry.objects.filter(
            tube_item=item, entry_type="production", date__year=year, date__month=month
        ).aggregate(t=Sum("quantity"))["t"] or 0
        sale = CycleTubeEntry.objects.filter(
            tube_item=item, entry_type="sale", date__year=year, date__month=month
        ).aggregate(t=Sum("quantity"))["t"] or 0

        if production or sale:
            rows.append({
                "item": item,
                "monthly_production": production,
                "monthly_sale": sale,
            })
            total_production += production
            total_sale += sale
            total_stock += item.stock
            grand_total += item.total_stock

    month_names = [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]
    month_name = f"{month_names[month-1]} {year}"

    context = {
        "rows": rows,
        "month_value": f"{year:04d}-{month:02d}",
        "month_name": month_name,
        "total_production": total_production,
        "total_sale": total_sale,
        "total_stock": total_stock,
        "grand_total": grand_total,
        "net_balance": total_production - total_sale,
    }
    return render(request, "cycletube/monthly_report.html", context)


# =====================================================================
# CYCLE TUBE PRODUCTION SUMMARY  (matches Excel sheet logic)
# =====================================================================
@login_required
def production_summary(request):
    """
    Daily production summary page for Cycle Tube.

    Auto-calculated columns:
      target_wt       = SUM(tube.weight * qty) per date  [inc VB]
      actual_wt_net   = actual_wt_gross * (1 - 0.0075)   [Incl VB, not Incl pack]
      variance_wt     = actual_wt_gross - target_wt
      target_consmpt  = target_wt * 0.015                [less VB]
      actual_comp_net = actual_wt_gross * 0.0225         [Weight (-Pck+VB)]
      variance_comp   = target_consmpt - actual_comp_net
      variance_mixing = target_wt - actual_mixing_compound
    """
    today = datetime.date.today()

    # ---- Handle POST (save manual entry) ----
    if request.method == "POST":
        form = CycleTubeDailyManualEntryForm(request.POST)
        if form.is_valid():
            entry_date = form.cleaned_data["date"]
            jali         = form.cleaned_data.get("jali") or Decimal("0.00")
            die_wastage  = form.cleaned_data.get("die_wastage") or Decimal("0.00")
            tube_cutting = form.cleaned_data.get("tube_cutting") or Decimal("0.00")
            total_waste  = form.cleaned_data.get("total_tube_waste") or Decimal("0.00")
            if total_waste <= 0:
                total_waste = jali + die_wastage + tube_cutting

            CycleTubeDailyManualEntry.objects.update_or_create(
                date=entry_date,
                defaults={
                    "valve_body_issued":      form.cleaned_data.get("valve_body_issued") or Decimal("0.00"),
                    "actual_wt_gross":        form.cleaned_data.get("actual_wt_gross") or Decimal("0.00"),
                    "actual_mixing_compound": form.cleaned_data.get("actual_mixing_compound") or Decimal("0.00"),
                    "jali":                   jali,
                    "die_wastage":            die_wastage,
                    "tube_cutting":           tube_cutting,
                    "total_tube_waste":       total_waste,
                }
            )
            messages.success(request, f"Manual entry for {entry_date.strftime('%d-%b-%Y')} saved.")
            from_d = request.POST.get("from_date", "")
            to_d   = request.POST.get("to_date", "")
            return redirect(f"{request.path}?from_date={from_d}&to_date={to_d}")

    # ---- Date range ----
    from_date_str  = request.GET.get("from_date", "").strip()
    to_date_str    = request.GET.get("to_date", "").strip()
    export_excel   = request.GET.get("export", "") == "excel"

    try:
        from_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d").date() if from_date_str else today.replace(day=1)
    except ValueError:
        from_date = today.replace(day=1)
    try:
        to_date = datetime.datetime.strptime(to_date_str, "%Y-%m-%d").date() if to_date_str else today
    except ValueError:
        to_date = today

    # ---- Aggregate production entries by date ----
    prod_entries = CycleTubeEntry.objects.filter(
        entry_type="production", date__gte=from_date, date__lte=to_date
    ).select_related("tube_item")

    by_date = {}
    for e in prod_entries:
        d = by_date.setdefault(e.date, {"production_pcs": 0, "target_wt": Decimal("0.00")})
        weight = e.tube_item.weight or Decimal("0.00")
        d["production_pcs"] += e.quantity
        d["target_wt"] += Decimal(str(e.quantity)) * weight

    # ---- Manual entries ----
    manual_entries = {
        m.date: m
        for m in CycleTubeDailyManualEntry.objects.filter(date__gte=from_date, date__lte=to_date)
    }

    all_dates = set(by_date.keys()) | set(manual_entries.keys())

    # ---- Build rows ----
    rows = []
    totals = {k: Decimal("0.00") for k in [
        "production_pcs", "valve_body_issued", "target_wt",
        "actual_wt_gross", "actual_wt_net", "variance_wt",
        "target_consmpt", "actual_comp_net", "variance_comp",
        "actual_mixing_compound", "variance_mixing",
        "jali", "die_wastage", "tube_cutting", "total_tube_waste",
    ]}
    totals["production_pcs"] = 0  # int

    for d in sorted(all_dates):
        stats  = by_date.get(d, {"production_pcs": 0, "target_wt": Decimal("0.00")})
        manual = manual_entries.get(d)

        pcs              = stats["production_pcs"]
        target_wt        = stats["target_wt"]
        valve_body       = manual.valve_body_issued      if manual else Decimal("0.00")
        actual_wt_gross  = manual.actual_wt_gross        if manual else Decimal("0.00")
        actual_mixing    = manual.actual_mixing_compound if manual else Decimal("0.00")
        jali             = manual.jali                   if manual else Decimal("0.00")
        die_wastage      = manual.die_wastage            if manual else Decimal("0.00")
        tube_cutting     = manual.tube_cutting           if manual else Decimal("0.00")
        total_tube_waste = manual.total_tube_waste       if manual else Decimal("0.00")

        pcs_dec          = Decimal(str(pcs))

        # ---- Auto-calculated columns (Exact Excel Formulas) ----
        # 1. Actual wt Net-less: Gross - (Pcs * 0.0075)
        actual_wt_net    = round(actual_wt_gross - (pcs_dec * PACK_FACTOR), 2)

        # 2. Variance Wt: Net-less - Target wt
        variance_wt      = round(actual_wt_net - target_wt, 2)

        # 3. Target Consmpt less VB: Target wt - (Pcs * 0.015)
        target_consmpt   = round(target_wt - (pcs_dec * VB_FACTOR), 2)

        # 4. Actual Comp Weight (-Pck+VB): Gross - (Pcs * 0.0225)
        actual_comp_net  = round(actual_wt_gross - (pcs_dec * COMB_FACTOR), 2)

        # 5. Variance Comp: Actual Comp - Target Consmpt
        variance_comp    = round(actual_comp_net - target_consmpt, 2)

        # 6. Variance Mixing: Actual Mixing - Target Consmpt
        variance_mixing  = round(actual_mixing - target_consmpt, 2)

        # 7. Total Tube Waste: Jali + Die Wastage + Tube Cutting (if manual total is 0)
        calculated_waste = jali + die_wastage + tube_cutting
        if total_tube_waste <= 0 and calculated_waste > 0:
            total_tube_waste = calculated_waste

        row = {
            "date": d,
            "production_pcs": pcs,
            "valve_body_issued": valve_body,
            "target_wt": round(target_wt, 2),
            "actual_wt_gross": actual_wt_gross,
            "actual_wt_net": actual_wt_net,
            "variance_wt": variance_wt,
            "target_consmpt": target_consmpt,
            "actual_comp_net": actual_comp_net,
            "variance_comp": variance_comp,
            "actual_mixing_compound": actual_mixing,
            "variance_mixing": variance_mixing,
            "jali": jali,
            "die_wastage": die_wastage,
            "tube_cutting": tube_cutting,
            "total_tube_waste": total_tube_waste,
        }
        rows.append(row)

        # Totals
        totals["production_pcs"]       += pcs
        totals["valve_body_issued"]    += valve_body
        totals["target_wt"]            += row["target_wt"]
        totals["actual_wt_gross"]      += actual_wt_gross
        totals["actual_wt_net"]        += actual_wt_net
        totals["variance_wt"]          += variance_wt
        totals["target_consmpt"]       += target_consmpt
        totals["actual_comp_net"]      += actual_comp_net
        totals["variance_comp"]        += variance_comp
        totals["actual_mixing_compound"] += actual_mixing
        totals["variance_mixing"]      += variance_mixing
        totals["jali"]                 += jali
        totals["die_wastage"]          += die_wastage
        totals["tube_cutting"]         += tube_cutting
        totals["total_tube_waste"]     += total_tube_waste

    # ---- Excel Export ----
    if export_excel:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cycle Tube Daily Summary"

        headers = [
            "Dated", "Prod Pcs", "Valve Body", "Target wt (Kgs)",
            "Actual wt Gross", "Actual wt Net-less", "Variance (Wt)",
            "Target Consmpt", "Actual Comp (-Pck+VB)", "Variance (Comp)",
            "Actual Mixing Compound", "Variance (Mixing)",
            "Jali", "Die Wastage", "Tube Cutting", "Total Tube Waste",
        ]
        ws.append(headers)
        hdr_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for r in rows:
            ws.append([
                r["date"].strftime("%d-%b-%Y"),
                r["production_pcs"],
                float(r["valve_body_issued"]),
                float(r["target_wt"]),
                float(r["actual_wt_gross"]),
                float(r["actual_wt_net"]),
                float(r["variance_wt"]),
                float(r["target_consmpt"]),
                float(r["actual_comp_net"]),
                float(r["variance_comp"]),
                float(r["actual_mixing_compound"]),
                float(r["variance_mixing"]),
                float(r["jali"]),
                float(r["die_wastage"]),
                float(r["tube_cutting"]),
                float(r["total_tube_waste"]),
            ])

        # Totals row
        tot_row = [
            "TOTAL",
            totals["production_pcs"],
            float(totals["valve_body_issued"]),
            float(totals["target_wt"]),
            float(totals["actual_wt_gross"]),
            float(totals["actual_wt_net"]),
            float(totals["variance_wt"]),
            float(totals["target_consmpt"]),
            float(totals["actual_comp_net"]),
            float(totals["variance_comp"]),
            float(totals["actual_mixing_compound"]),
            float(totals["variance_mixing"]),
            float(totals["jali"]),
            float(totals["die_wastage"]),
            float(totals["tube_cutting"]),
            float(totals["total_tube_waste"]),
        ]
        ws.append(tot_row)
        last_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=last_row, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 16

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = (
            f'attachment; filename="CycleTube_Summary_{from_date}_to_{to_date}.xlsx"'
        )
        wb.save(response)
        return response

    manual_form = CycleTubeDailyManualEntryForm(initial={"date": today})

    context = {
        "rows": rows,
        "totals": totals,
        "from_date": from_date.strftime("%Y-%m-%d"),
        "to_date": to_date.strftime("%Y-%m-%d"),
        "manual_form": manual_form,
    }
    return render(request, "cycletube/production_summary.html", context)