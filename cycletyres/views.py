import datetime
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Q
from django.shortcuts import render, redirect

from .models import CycleTyreItem, CycleTyreEntry, BUCKET_CHOICES
from .forms import CycleTyreItemForm, ProductionEntryForm, SaleEntryForm, AdjustmentEntryForm


@login_required
def dashboard(request):
    q = request.GET.get("q", "").strip()
    items = CycleTyreItem.objects.filter(is_active=True)
    if q:
        items = items.filter(
            Q(size__icontains=q) | Q(box_type__icontains=q) | Q(material__icontains=q) | Q(brand__icontains=q)
        )

    today = datetime.date.today()
    today_production = CycleTyreEntry.objects.filter(entry_type="production", date=today).aggregate(t=Sum("quantity"))["t"] or 0
    today_sale = CycleTyreEntry.objects.filter(entry_type="sale", date=today).aggregate(t=Sum("quantity"))["t"] or 0

    month_start = today.replace(day=1)
    month_production = CycleTyreEntry.objects.filter(entry_type="production", date__gte=month_start, date__lte=today).aggregate(t=Sum("quantity"))["t"] or 0
    month_sale = CycleTyreEntry.objects.filter(entry_type="sale", date__gte=month_start, date__lte=today).aggregate(t=Sum("quantity"))["t"] or 0

    for item in items:
        item.month_production = CycleTyreEntry.objects.filter(
            tyre_item=item, entry_type="production", date__gte=month_start, date__lte=today
        ).aggregate(t=Sum("quantity"))["t"] or 0
        item.month_sale = CycleTyreEntry.objects.filter(
            tyre_item=item, entry_type="sale", date__gte=month_start, date__lte=today
        ).aggregate(t=Sum("quantity"))["t"] or 0

    grand_total = sum(item.total_stock for item in items)
    total_stock_col = sum(item.stock for item in items)
    total_second_col = sum(item.second_stock for item in items)
    total_rfm_col = sum(item.rfm_stock for item in items)

    context = {
        "items": items,
        "q": q,
        "today": today,
        "today_production": today_production,
        "today_sale": today_sale,
        "month_production": month_production,
        "month_sale": month_sale,
        "grand_total": grand_total,
        "total_stock_col": total_stock_col,
        "total_second_col": total_second_col,
        "total_rfm_col": total_rfm_col,
    }
    return render(request, "cycletyres/dashboard.html", context)


@login_required
def add_item(request):
    form = CycleTyreItemForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.box_type = item.box_type.strip()
        item.size = item.size.strip()
        item.material = item.material.strip()
        item.brand = item.brand.strip()
        item.save()
        messages.success(request, f"Cycle Tyre '{item}' add ho gaya.")
        return redirect("cycletyre_dashboard")
    return render(request, "cycletyres/add_item.html", {"form": form})


@login_required
def add_production(request):
    form = ProductionEntryForm(request.POST or None, initial={"date": datetime.date.today()})
    if request.method == "POST" and form.is_valid():
        item = form.cleaned_data["tyre_item"]
        all_curing = form.cleaned_data["all_curing"]
        second_grade = form.cleaned_data.get("second_grade", 0) or 0
        rejected_grade = form.cleaned_data.get("rejected_grade", 0) or 0
        
        first_grade = all_curing - (second_grade + rejected_grade)
        
        if first_grade < 0:
            messages.error(
                request,
                f"Invalid entry: 2nd Grade ({second_grade}) + Rejected ({rejected_grade}) "
                f"is greater than All Curing ({all_curing})!"
            )
        else:
            with transaction.atomic():
                # 1st Grade (Black) goes to main stock, 2nd Grade goes to second_stock
                item.stock += first_grade
                item.second_stock += second_grade
                item.save(update_fields=["stock", "second_stock"])
                
                CycleTyreEntry.objects.create(
                    tyre_item=item,
                    entry_type="production",
                    bucket="stock",
                    quantity=first_grade,
                    all_curing=all_curing,
                    first_grade=first_grade,
                    second_grade=second_grade,
                    rejected_grade=rejected_grade,
                    date=form.cleaned_data["date"],
                    remark=form.cleaned_data.get("remark", ""),
                    user=request.user,
                )
            messages.success(
                request,
                f"Production Entry Saved: {item} | All Curing: {all_curing} -> "
                f"1st Grade (Black): +{first_grade}, 2nd Grade: +{second_grade}, Rejected: {rejected_grade}"
            )
            return redirect("cycletyre_add_production")
            
    recent = CycleTyreEntry.objects.filter(entry_type="production").select_related("tyre_item", "user")[:15]
    return render(request, "cycletyres/add_production.html", {"form": form, "recent": recent})


@login_required
def second_grade_stock(request):
    q = request.GET.get("q", "").strip()
    items = CycleTyreItem.objects.filter(is_active=True)
    if q:
        items = items.filter(
            Q(size__icontains=q) | Q(box_type__icontains=q) | Q(material__icontains=q) | Q(brand__icontains=q)
        )

    total_second_stock = sum(item.second_stock for item in items)
    
    # Recent 2nd grade production logs
    second_grade_entries = CycleTyreEntry.objects.filter(
        entry_type="production", second_grade__gt=0
    ).select_related("tyre_item", "user")[:25]

    context = {
        "items": items,
        "q": q,
        "total_second_stock": total_second_stock,
        "second_grade_entries": second_grade_entries,
    }
    return render(request, "cycletyres/second_grade_stock.html", context)


@login_required
def add_sale(request):
    form = SaleEntryForm(request.POST or None, initial={"date": datetime.date.today()})
    if request.method == "POST" and form.is_valid():
        item = form.cleaned_data["tyre_item"]
        bucket = form.cleaned_data["bucket"]
        qty = form.cleaned_data["quantity"]
        bill_number = form.cleaned_data["bill_number"].strip()
        current = getattr(item, bucket)

        duplicate = CycleTyreEntry.objects.filter(
            entry_type="sale", bill_number__iexact=bill_number
        ).first()

        if duplicate:
            messages.error(
                request,
                f"Bill number '{bill_number}' pehle se use ho chuka hai "
                f"({duplicate.date} - {duplicate.tyre_item}). Dubara save nahi hoga, "
                f"bill number check kar lo."
            )
        elif qty > current:
            messages.error(request, f"'{item}' ke '{dict(BUCKET_CHOICES)[bucket]}' bucket mai sirf {current} hi available hai.")
        else:
            with transaction.atomic():
                setattr(item, bucket, current - qty)
                item.save(update_fields=[bucket])
                CycleTyreEntry.objects.create(
                    tyre_item=item,
                    entry_type="sale",
                    bucket=bucket,
                    quantity=qty,
                    date=form.cleaned_data["date"],
                    bill_number=bill_number,
                    remark=form.cleaned_data["remark"],
                    user=request.user,
                )
            messages.success(request, f"Sale entry ho gayi: {item} -{qty} ({dict(BUCKET_CHOICES)[bucket]}) | Bill: {bill_number}")
            return redirect("cycletyre_add_sale")
    recent = CycleTyreEntry.objects.filter(entry_type="sale").select_related("tyre_item", "user")[:15]
    return render(request, "cycletyres/add_sale.html", {"form": form, "recent": recent})


@login_required
def add_adjustment(request):
    form = AdjustmentEntryForm(request.POST or None, initial={"date": datetime.date.today()})
    if request.method == "POST" and form.is_valid():
        item = form.cleaned_data["tyre_item"]
        bucket = form.cleaned_data["bucket"]
        action = form.cleaned_data["action"]
        qty = form.cleaned_data["quantity"]
        current = getattr(item, bucket)
        if action == "subtract" and qty > current:
            messages.error(request, f"'{item}' ke '{dict(BUCKET_CHOICES)[bucket]}' bucket mai sirf {current} hi available hai.")
        else:
            signed_qty = qty if action == "add" else -qty
            with transaction.atomic():
                setattr(item, bucket, current + signed_qty)
                item.save(update_fields=[bucket])
                CycleTyreEntry.objects.create(
                    tyre_item=item,
                    entry_type="adjustment",
                    bucket=bucket,
                    quantity=signed_qty,
                    date=form.cleaned_data["date"],
                    remark=form.cleaned_data["remark"],
                    user=request.user,
                )
            messages.success(request, f"Adjustment ho gaya: {item} {signed_qty:+d} ({dict(BUCKET_CHOICES)[bucket]})")
            return redirect("cycletyre_add_adjustment")
    return render(request, "cycletyres/add_adjustment.html", {"form": form})


@login_required
def entries_log(request):
    date_str = request.GET.get("date", "")
    month_str = request.GET.get("month", "")
    entry_type = request.GET.get("type", "")

    entries = CycleTyreEntry.objects.select_related("tyre_item", "user")

    if date_str:
        try:
            d = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            entries = entries.filter(date=d)
        except ValueError:
            pass
    elif month_str:
        try:
            y, m = month_str.split("-")
            entries = entries.filter(date__year=int(y), date__month=int(m))
        except ValueError:
            pass

    if entry_type:
        entries = entries.filter(entry_type=entry_type)

    entries = entries[:500]

    context = {
        "entries": entries,
        "date_str": date_str,
        "month_str": month_str,
        "entry_type": entry_type,
    }
    return render(request, "cycletyres/entries_log.html", context)


@login_required
def monthly_report(request):
    month_str = request.GET.get("month", "")
    today = datetime.date.today()
    if month_str:
        try:
            y, m = month_str.split("-")
            year, month = int(y), int(m)
        except ValueError:
            year, month = today.year, today.month
    else:
        year, month = today.year, today.month

    items = CycleTyreItem.objects.filter(is_active=True)
    rows = []
    for item in items:
        production = CycleTyreEntry.objects.filter(
            tyre_item=item, entry_type="production", date__year=year, date__month=month
        ).aggregate(t=Sum("quantity"))["t"] or 0
        sale = CycleTyreEntry.objects.filter(
            tyre_item=item, entry_type="sale", date__year=year, date__month=month
        ).aggregate(t=Sum("quantity"))["t"] or 0
        if production or sale:
            rows.append({
                "item": item,
                "monthly_production": production,
                "monthly_sale": sale,
            })

    context = {
        "rows": rows,
        "month_value": f"{year:04d}-{month:02d}",
    }
    return render(request, "cycletyres/monthly_report.html", context)


from decimal import Decimal, InvalidOperation
from django.http import HttpResponse

COMPOUND_PERCENT = Decimal("0.825")  # Theoretical Total Compound = Theoretical KG x 82.5%


@login_required
def daily_summary(request):
    from .models import CycleTyreDailyManualEntry
    from .forms import CycleTyreDailyManualEntryForm
    
    today = datetime.date.today()

    if request.method == "POST":
        form = CycleTyreDailyManualEntryForm(request.POST)
        if form.is_valid():
            entry_date = form.cleaned_data["date"]
            CycleTyreDailyManualEntry.objects.update_or_create(
                date=entry_date,
                defaults={
                    "parchi_kg": form.cleaned_data.get("parchi_kg") or Decimal("0.00"),
                    "mixing_actual_compound": form.cleaned_data.get("mixing_actual_compound") or Decimal("0.00"),
                    "chakka": form.cleaned_data.get("chakka") or Decimal("0.00"),
                    "calander_bias_cutt": form.cleaned_data.get("calander_bias_cutt") or Decimal("0.00"),
                    "packing_wastage": form.cleaned_data.get("packing_wastage") or Decimal("0.00"),
                    "tar": form.cleaned_data.get("tar") or Decimal("0.00"),
                }
            )
            messages.success(request, f"Daily summary manual entry for {entry_date.strftime('%d-%b-%Y')} saved.")
            from_d = request.POST.get("from_date", "")
            to_d = request.POST.get("to_date", "")
            return redirect(f"{request.path}?from_date={from_d}&to_date={to_d}")

    from_date_str = request.GET.get("from_date", "").strip()
    to_date_str = request.GET.get("to_date", "").strip()
    export_excel = request.GET.get("export", "") == "excel"

    try:
        from_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d").date() if from_date_str else today.replace(day=1)
    except ValueError:
        from_date = today.replace(day=1)
    try:
        to_date = datetime.datetime.strptime(to_date_str, "%Y-%m-%d").date() if to_date_str else today
    except ValueError:
        to_date = today

    prod_entries = CycleTyreEntry.objects.filter(
        entry_type="production", date__gte=from_date, date__lte=to_date
    ).select_related("tyre_item")

    by_date = {}
    for e in prod_entries:
        d = by_date.setdefault(e.date, {"production_pcs": 0, "packing_pcs": 0, "theoretical_kg": Decimal("0.00")})
        curing_qty = e.all_curing if e.all_curing > 0 else e.quantity
        first_grade_qty = e.first_grade if e.first_grade > 0 else e.quantity
        weight = e.tyre_item.weight or Decimal("0.00")

        d["production_pcs"] += curing_qty
        d["packing_pcs"] += first_grade_qty
        d["theoretical_kg"] += Decimal(str(curing_qty)) * weight

    manual_entries = {
        m.date: m for m in CycleTyreDailyManualEntry.objects.filter(date__gte=from_date, date__lte=to_date)
    }

    all_dates = set(by_date.keys()).union(set(manual_entries.keys()))

    rows = []
    totals = {
        "production_pcs": 0,
        "packing_pcs": 0,
        "theoretical_kg": Decimal("0.00"),
        "parchi_kg": Decimal("0.00"),
        "difference": Decimal("0.00"),
        "theoretical_total_compound": Decimal("0.00"),
        "mixing_actual_compound": Decimal("0.00"),
        "variance": Decimal("0.00"),
        "chakka": Decimal("0.00"),
        "calander_bias_cutt": Decimal("0.00"),
        "packing_wastage": Decimal("0.00"),
        "tar": Decimal("0.00"),
    }

    for d in sorted(all_dates, reverse=True):
        stats = by_date.get(d, {"production_pcs": 0, "packing_pcs": 0, "theoretical_kg": Decimal("0.00")})
        manual = manual_entries.get(d)

        parchi_kg = manual.parchi_kg if manual else Decimal("0.00")
        mixing_actual = manual.mixing_actual_compound if manual else Decimal("0.00")
        chakka = manual.chakka if manual else Decimal("0.00")
        calander = manual.calander_bias_cutt if manual else Decimal("0.00")
        packing_w = manual.packing_wastage if manual else Decimal("0.00")
        tar = manual.tar if manual else Decimal("0.00")

        theoretical_kg = stats["theoretical_kg"]
        difference = parchi_kg - theoretical_kg
        theoretical_total_compound = theoretical_kg * COMPOUND_PERCENT
        variance = mixing_actual - theoretical_total_compound

        row = {
            "date": d,
            "particulars": "Cycle Tyre",
            "production_pcs": stats["production_pcs"],
            "packing_pcs": stats["packing_pcs"],
            "theoretical_kg": round(theoretical_kg, 2),
            "parchi_kg": parchi_kg,
            "difference": round(difference, 2),
            "theoretical_total_compound": round(theoretical_total_compound, 2),
            "mixing_actual_compound": mixing_actual,
            "variance": round(variance, 2),
            "chakka": chakka,
            "calander_bias_cutt": calander,
            "packing_wastage": packing_w,
            "tar": tar,
        }
        rows.append(row)

        totals["production_pcs"] += row["production_pcs"]
        totals["packing_pcs"] += row["packing_pcs"]
        totals["theoretical_kg"] += row["theoretical_kg"]
        totals["parchi_kg"] += row["parchi_kg"]
        totals["difference"] += row["difference"]
        totals["theoretical_total_compound"] += row["theoretical_total_compound"]
        totals["mixing_actual_compound"] += row["mixing_actual_compound"]
        totals["variance"] += row["variance"]
        totals["chakka"] += row["chakka"]
        totals["calander_bias_cutt"] += row["calander_bias_cutt"]
        totals["packing_wastage"] += row["packing_wastage"]
        totals["tar"] += row["tar"]

    if export_excel:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cycle Tyre Daily Summary"

        headers = [
            "Dated", "Particulars", "PRODUCTION PCS", "Packing", "Theoretical KG",
            "Packing Parch Kg", "Difference", "Theoretical TOTAL COMPOUND",
            "MIXING ACTUAL COMPOUND", "Variance", "Chakka", "Calander Bias Cutt.",
            "Packing Wastage", "Tar"
        ]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

        for r in rows:
            ws.append([
                r["date"].strftime("%d-%b-%Y"),
                r["particulars"],
                r["production_pcs"],
                r["packing_pcs"],
                float(r["theoretical_kg"]),
                float(r["parchi_kg"]),
                float(r["difference"]),
                float(r["theoretical_total_compound"]),
                float(r["mixing_actual_compound"]),
                float(r["variance"]),
                float(r["chakka"]),
                float(r["calander_bias_cutt"]),
                float(r["packing_wastage"]),
                float(r["tar"]),
            ])

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="Cycle_Tyre_Daily_Summary_{from_date}_to_{to_date}.xlsx"'
        wb.save(response)
        return response

    manual_form = CycleTyreDailyManualEntryForm(initial={"date": today})

    context = {
        "rows": rows,
        "totals": totals,
        "from_date": from_date.strftime("%Y-%m-%d"),
        "to_date": to_date.strftime("%Y-%m-%d"),
        "manual_form": manual_form,
    }
    return render(request, "cycletyres/daily_summary.html", context)