import datetime
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Q
from django.shortcuts import render, redirect

from .models import TyreItem, DailyEntry, BUCKET_CHOICES
from .forms import TyreItemForm, ProductionEntryForm, DispatchEntryForm, AdjustmentEntryForm


def login_view(request):
    if request.user.is_authenticated:
        return redirect("home")
    form = AuthenticationForm(request, data=request.POST or None)
    if request.method == "POST" and form.is_valid():
        auth_login(request, form.get_user())
        return redirect("home")
    return render(request, "stock/login.html", {"form": form})



@login_required
def home(request):
    return render(request, "home.html")

def logout_view(request):
    auth_logout(request)
    return redirect("login")


@login_required
def dashboard(request):
    q = request.GET.get("q", "").strip()
    items = TyreItem.objects.filter(is_active=True)
    if q:
        items = items.filter(
            Q(tyre__icontains=q) | Q(pattern__icontains=q) | Q(type__icontains=q)
        )

    today = datetime.date.today()
    today_production = DailyEntry.objects.filter(entry_type="production", date=today).aggregate(t=Sum("quantity"))["t"] or 0
    today_dispatch = DailyEntry.objects.filter(entry_type="dispatch", date=today).aggregate(t=Sum("quantity"))["t"] or 0

    month_start = today.replace(day=1)
    
    for item in items:
        item.month_curing = DailyEntry.objects.filter(
            tyre_item=item, entry_type="production",
            date__gte=month_start, date__lte=today
        ).aggregate(t=Sum("quantity"))["t"] or 0
        item.month_despatch = DailyEntry.objects.filter(
            tyre_item=item, entry_type="dispatch",
            date__gte=month_start, date__lte=today
        ).aggregate(t=Sum("quantity"))["t"] or 0

   
    month_production = DailyEntry.objects.filter(entry_type="production", date__gte=month_start, date__lte=today).aggregate(t=Sum("quantity"))["t"] or 0
    month_dispatch = DailyEntry.objects.filter(entry_type="dispatch", date__gte=month_start, date__lte=today).aggregate(t=Sum("quantity"))["t"] or 0

    grand_total = sum(item.total_stock for item in items)
    total_stock_col = sum(item.stock for item in items)
    total_repair = sum(item.repair_tyre_stock for item in items)
    total_rfm = sum(item.rfm_ok_tyre for item in items)
    total_old = sum(item.old_tyres_2025 for item in items)
    total_hold = sum(item.on_hold_export for item in items)
    total_curing = sum(item.month_curing for item in items)
    total_despatch = sum(item.month_despatch for item in items)
    context = {
        "items": items,
        "q": q,
        "today": today,
        "today_production": today_production,
        "today_dispatch": today_dispatch,
        "month_production": month_production,
        "month_dispatch": month_dispatch,
        "grand_total": grand_total,
        "total_stock_col": total_stock_col,
        "total_repair": total_repair,
        "total_rfm": total_rfm,
        "total_old": total_old,
        "total_hold": total_hold,
        "total_curing": total_curing,
        "total_despatch": total_despatch,
    }
    return render(request, "stock/dashboard.html", context)


@login_required
def add_tyre(request):
    form = TyreItemForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.tyre = item.tyre.strip()
        item.pattern = item.pattern.strip()
        item.type = item.type.strip()
        item.save()
        messages.success(request, f"Tyre '{item}' add ho gaya.")
        return redirect("dashboard")
    return render(request, "stock/add_tyre.html", {"form": form})


@login_required
def add_production(request):
    form = ProductionEntryForm(request.POST or None, initial={"date": datetime.date.today()})
    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            item = form.cleaned_data["tyre_item"]
            packing = form.cleaned_data["packing"]
            item.stock += packing
            item.save(update_fields=["stock"])
            DailyEntry.objects.create(
                tyre_item=item,
                entry_type="production",
                bucket="stock",
                quantity=packing,
                date=form.cleaned_data["date"],
                all_curing=form.cleaned_data["all_curing"],
                production_tyre=form.cleaned_data["production_tyre"],
                repair=form.cleaned_data["repair"],
                second_grade=form.cleaned_data["second_grade"],
                third_grade=form.cleaned_data["third_grade"],
                lose_tyre=form.cleaned_data["lose_tyre"],
                actual_weight=form.cleaned_data["actual_weight"],
                remark=form.cleaned_data["remark"],
                user=request.user,
            )
        messages.success(request, f"Production entry ho gayi: {item} | Packing (Stock mein add) +{packing}")
        return redirect("add_production")
    recent = DailyEntry.objects.filter(entry_type="production").select_related("tyre_item", "user")[:15]
    return render(request, "stock/add_production.html", {"form": form, "recent": recent})

@login_required
def add_dispatch(request):
    form = DispatchEntryForm(request.POST or None, initial={"date": datetime.date.today()})
    if request.method == "POST" and form.is_valid():
            item = form.cleaned_data["tyre_item"]
            bucket = form.cleaned_data["bucket"]
            qty = form.cleaned_data["quantity"]
            bill_number = form.cleaned_data["bill_number"].strip()
            current = getattr(item, bucket)

            duplicate = DailyEntry.objects.filter(
                entry_type="dispatch", bill_number__iexact=bill_number
            ).first()

            if duplicate:
                messages.error(
                    request,
                    f"Bill number '{bill_number}' pehle se use ho chuka hai "
                    f"({duplicate.date} - {duplicate.tyre_item}). Dubara save nahi hoga, "
                    f"bill number check kar lo."
                )
            elif qty > current:
                messages.error(request, f"'{item}' ke '{dict(BUCKET_CHOICES)[bucket]}' bucket mai sirf {current} hi available hai.")
            else:
                with transaction.atomic():
                    setattr(item, bucket, current - qty)
                    item.save(update_fields=[bucket])
                    DailyEntry.objects.create(
                        tyre_item=item,
                        entry_type="dispatch",
                        bucket=bucket,
                        quantity=qty,
                        date=form.cleaned_data["date"],
                        bill_number=bill_number,
                        remark=form.cleaned_data["remark"],
                        user=request.user,
                    )
                messages.success(request, f"Dispatch entry has been added: {item} -{qty} ({dict(BUCKET_CHOICES)[bucket]}) | Bill: {bill_number}")
                return redirect("add_dispatch")
    recent = DailyEntry.objects.filter(entry_type="dispatch").select_related("tyre_item", "user")[:15]
    return render(request, "stock/add_dispatch.html", {"form": form, "recent": recent})


@login_required
def add_adjustment(request):
    form = AdjustmentEntryForm(request.POST or None, initial={"date": datetime.date.today()})
    if request.method == "POST" and form.is_valid():
        item = form.cleaned_data["tyre_item"]
        bucket = form.cleaned_data["bucket"]
        action = form.cleaned_data["action"]
        qty = form.cleaned_data["quantity"]
        current = getattr(item, bucket)
        if action == "subtract" and qty > current:
            messages.error(request, f"'{item}' ke '{dict(BUCKET_CHOICES)[bucket]}' bucket mai sirf {current} hi available hai.")
        else:
            signed_qty = qty if action == "add" else -qty
            with transaction.atomic():
                setattr(item, bucket, current + signed_qty)
                item.save(update_fields=[bucket])
                DailyEntry.objects.create(
                    tyre_item=item,
                    entry_type="adjustment",
                    bucket=bucket,
                    quantity=signed_qty,
                    date=form.cleaned_data["date"],
                    remark=form.cleaned_data["remark"],
                    user=request.user,
                )
            messages.success(request, f"Adjustment ho gaya: {item} {signed_qty:+d} ({dict(BUCKET_CHOICES)[bucket]})")
            return redirect("add_adjustment")
    return render(request, "stock/add_adjustment.html", {"form": form})


@login_required
def entries_log(request):
    date_str = request.GET.get("date", "")
    month_str = request.GET.get("month", "")
    entry_type = request.GET.get("type", "")

    entries = DailyEntry.objects.select_related("tyre_item", "user")

    if date_str:
        try:
            d = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            entries = entries.filter(date=d)
        except ValueError:
            pass
    elif month_str:
        try:
            y, m = month_str.split("-")
            entries = entries.filter(date__year=int(y), date__month=int(m))
        except ValueError:
            pass

    if entry_type:
        entries = entries.filter(entry_type=entry_type)

    entries = entries[:500]

    context = {
        "entries": entries,
        "date_str": date_str,
        "month_str": month_str,
        "entry_type": entry_type,
    }
    return render(request, "stock/entries_log.html", context)


@login_required
def monthly_report(request):
    month_str = request.GET.get("month", "")
    today = datetime.date.today()
    if month_str:
        try:
            y, m = month_str.split("-")
            year, month = int(y), int(m)
        except ValueError:
            year, month = today.year, today.month
    else:
        year, month = today.year, today.month

    items = TyreItem.objects.filter(is_active=True)
    rows = []
    for item in items:
        production = DailyEntry.objects.filter(
            tyre_item=item, entry_type="production", date__year=year, date__month=month
        ).aggregate(t=Sum("quantity"))["t"] or 0
        dispatch = DailyEntry.objects.filter(
            tyre_item=item, entry_type="dispatch", date__year=year, date__month=month
        ).aggregate(t=Sum("quantity"))["t"] or 0
        if production or dispatch:
            rows.append({
                "item": item,
                "monthly_curing": production,
                "monthly_despatch": dispatch,
            })

    total_curing = sum(r["monthly_curing"] for r in rows)
    total_despatch = sum(r["monthly_despatch"] for r in rows)
    total_stock_col = sum(item.stock for item in items)
    grand_total = sum(item.total_stock for item in items)
    net_balance = total_curing - total_despatch

    selected_month_date = datetime.date(year, month, 1)
    month_name = selected_month_date.strftime("%B %Y")

    # Real 6-month trend (including the selected month), not fake/random data
    trend_labels, trend_production, trend_dispatch = [], [], []
    cursor_year, cursor_month = year, month
    months_back = []
    for _ in range(6):
        months_back.append((cursor_year, cursor_month))
        cursor_month -= 1
        if cursor_month == 0:
            cursor_month = 12
            cursor_year -= 1
    months_back.reverse()

    for y2, m2 in months_back:
        prod = DailyEntry.objects.filter(
            entry_type="production", date__year=y2, date__month=m2
        ).aggregate(t=Sum("quantity"))["t"] or 0
        disp = DailyEntry.objects.filter(
            entry_type="dispatch", date__year=y2, date__month=m2
        ).aggregate(t=Sum("quantity"))["t"] or 0
        trend_labels.append(datetime.date(y2, m2, 1).strftime("%b %Y"))
        trend_production.append(prod)
        trend_dispatch.append(disp)

    context = {
        "rows": rows,
        "month_value": f"{year:04d}-{month:02d}",
        "month_name": month_name,
        "total_curing": total_curing,
        "total_despatch": total_despatch,
        "total_stock_col": total_stock_col,
        "grand_total": grand_total,
        "net_balance": net_balance,
        "trend_labels": trend_labels,
        "trend_production": trend_production,
        "trend_dispatch": trend_dispatch,
    }
    return render(request, "stock/monthly_report.html", context)



from django.db.models import Sum, Max, Q
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Max, Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import HttpResponse
from datetime import date, datetime, timedelta

import datetime

@login_required
def production_sheet(request):
    date_str = request.GET.get("date", "")
    month_str = request.GET.get("month", "")
    export_excel = request.GET.get("export") == "excel"
    page = request.GET.get("page", 1)
    today = datetime.date.today()

    entries = DailyEntry.objects.filter(entry_type="production").select_related("tyre_item")

    # Determine date range
    if date_str:
        try:
            d = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            entries = entries.filter(date=d)
            range_label = d.strftime("%d %b %Y")
            date_str = d.strftime("%Y-%m-%d")
        except ValueError:
            date_str = ""
            month_str = f"{today.year:04d}-{today.month:02d}"
            entries = entries.filter(date__year=today.year, date__month=today.month)
            range_label = today.strftime("%B %Y")
    else:
        if month_str:
            try:
                y, m = month_str.split("-")
                year, month = int(y), int(m)
            except ValueError:
                year, month = today.year, today.month
        else:
            year, month = today.year, today.month
        month_str = f"{year:04d}-{month:02d}"
        entries = entries.filter(date__year=year, date__month=month)
        range_label = datetime.date(year, month, 1).strftime("%B %Y")

    entries = entries.order_by("-date", "tyre_item__tyre")

    # Next/Prev date navigation
    if date_str:
        current_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        prev_date = current_date - datetime.timedelta(days=1)
        next_date = current_date + datetime.timedelta(days=1)
    else:
        year, month = map(int, month_str.split("-"))
        first_day = datetime.date(year, month, 1)
        if month == 12:
            last_day = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
        else:
            last_day = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
        prev_date = first_day - datetime.timedelta(days=1)
        next_date = last_day + datetime.timedelta(days=1)

    # Aggregation logic
    if date_str:
        # DATE VIEW – keep individual rows
        paginated_entries = entries
    else:
        # MONTH VIEW – aggregate by tyre
        entries = entries.values("tyre_item").annotate(
            total_all_curing=Sum("all_curing"),
            total_production_tyre=Sum("production_tyre"),
            total_repair=Sum("repair"),
            total_second_grade=Sum("second_grade"),
            total_third_grade=Sum("third_grade"),
            total_lose_tyre=Sum("lose_tyre"),
            total_packing=Sum("quantity"),
            latest_date=Max("date"),
        ).order_by("-latest_date")

    # Pagination (only for DATE VIEW)
    if date_str:
        paginator = Paginator(entries, 50)
        try:
            page_obj = paginator.page(page)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
    else:
        page_obj = entries

    # RFM adjustments (only for date view)
    if date_str:
        date_list = list(entries.values_list("date", flat=True).distinct())
        rfm_qs = DailyEntry.objects.filter(
            entry_type="adjustment", bucket="rfm_ok_tyre", date__in=date_list
        ).values("date", "tyre_item_id").annotate(total=Sum("quantity"))
        rfm_map = {(r["date"], r["tyre_item_id"]): r["total"] for r in rfm_qs}

    rows = []
    totals = {
        "all_curing": 0, "production_tyre": 0, "repair": 0,
        "second_grade": 0, "third_grade": 0, "lose_tyre": 0,
        "packing": 0, "rfm": 0,
    }
    daily_packing = {}
    daily_repair = {}
    daily_second = {}
    daily_third = {}
    daily_lose = {}
    daily_curing = {}

    if date_str:
        # DATE VIEW
        for e in page_obj:
            rfm_val = rfm_map.get((e.date, e.tyre_item_id), 0)
            rows.append({
                "entry": e,
                "rfm": rfm_val,
                "is_date_view": True,
            })
            totals["all_curing"] += e.all_curing
            totals["production_tyre"] += e.production_tyre
            totals["repair"] += e.repair
            totals["second_grade"] += e.second_grade
            totals["third_grade"] += e.third_grade
            totals["lose_tyre"] += e.lose_tyre
            totals["packing"] += e.quantity
            totals["rfm"] += rfm_val
            
            daily_packing[e.date] = daily_packing.get(e.date, 0) + e.quantity
            daily_repair[e.date] = daily_repair.get(e.date, 0) + e.repair
            daily_second[e.date] = daily_second.get(e.date, 0) + e.second_grade
            daily_third[e.date] = daily_third.get(e.date, 0) + e.third_grade
            daily_lose[e.date] = daily_lose.get(e.date, 0) + e.lose_tyre
            daily_curing[e.date] = daily_curing.get(e.date, 0) + e.all_curing
    else:
        # MONTH VIEW
        for row in page_obj:
            tyre_item = DailyEntry.objects.filter(tyre_item=row["tyre_item"]).first().tyre_item
            rows.append({
                "entry": tyre_item,
                "rfm": 0,
                "is_date_view": False,
                "aggregated": {
                    "all_curing": row["total_all_curing"],
                    "production_tyre": row["total_production_tyre"],
                    "repair": row["total_repair"],
                    "second_grade": row["total_second_grade"],
                    "third_grade": row["total_third_grade"],
                    "lose_tyre": row["total_lose_tyre"],
                    "packing": row["total_packing"],
                }
            })
            totals["all_curing"] += row["total_all_curing"]
            totals["production_tyre"] += row["total_production_tyre"]
            totals["repair"] += row["total_repair"]
            totals["second_grade"] += row["total_second_grade"]
            totals["third_grade"] += row["total_third_grade"]
            totals["lose_tyre"] += row["total_lose_tyre"]
            totals["packing"] += row["total_packing"]

    # ✅ FIX: Charts for BOTH date view and month view
    if date_str:
        # Date view – date-wise data
        chart_dates = sorted(daily_packing.keys())
        chart_labels = [d.strftime("%d-%b") for d in chart_dates]
        chart_values = [daily_packing[d] for d in chart_dates]
        chart_repair = [daily_repair[d] for d in chart_dates]
        chart_second = [daily_second[d] for d in chart_dates]
        chart_third = [daily_third[d] for d in chart_dates]
        chart_lose = [daily_lose[d] for d in chart_dates]
        chart_curing = [daily_curing[d] for d in chart_dates]
    else:
        # Month view – tyre-wise data (aggregated per tyre)
        chart_labels = [str(row["entry"]) for row in rows]  # Tyre names as labels
        chart_values = [row["aggregated"]["packing"] for row in rows]
        chart_repair = [row["aggregated"]["repair"] for row in rows]
        chart_second = [row["aggregated"]["second_grade"] for row in rows]
        chart_third = [row["aggregated"]["third_grade"] for row in rows]
        chart_lose = [row["aggregated"]["lose_tyre"] for row in rows]
        chart_curing = [row["aggregated"]["all_curing"] for row in rows]

    # Export to Excel
    if export_excel:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Production Sheet"

        headers = ["Date", "Tyre", "All Curing", "Production Tyre", "Repair", "2nd", "3rd", "Lose Tyre", "RFM", "Packing"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        for row_idx, row in enumerate(rows, 2):
            if row.get("is_date_view"):
                ws.cell(row=row_idx, column=1, value=row["entry"].date.strftime("%d-%b-%Y"))
                ws.cell(row=row_idx, column=2, value=str(row["entry"].tyre_item))
                ws.cell(row=row_idx, column=3, value=row["entry"].all_curing)
                ws.cell(row=row_idx, column=4, value=row["entry"].production_tyre)
                ws.cell(row=row_idx, column=5, value=row["entry"].repair)
                ws.cell(row=row_idx, column=6, value=row["entry"].second_grade)
                ws.cell(row=row_idx, column=7, value=row["entry"].third_grade)
                ws.cell(row=row_idx, column=8, value=row["entry"].lose_tyre)
                ws.cell(row=row_idx, column=9, value=row["rfm"])
                ws.cell(row=row_idx, column=10, value=row["entry"].quantity)
            else:
                ws.cell(row=row_idx, column=1, value="MONTH TOTAL")
                ws.cell(row=row_idx, column=2, value=str(row["entry"]))
                ws.cell(row=row_idx, column=3, value=row["aggregated"]["all_curing"])
                ws.cell(row=row_idx, column=4, value=row["aggregated"]["production_tyre"])
                ws.cell(row=row_idx, column=5, value=row["aggregated"]["repair"])
                ws.cell(row=row_idx, column=6, value=row["aggregated"]["second_grade"])
                ws.cell(row=row_idx, column=7, value=row["aggregated"]["third_grade"])
                ws.cell(row=row_idx, column=8, value=row["aggregated"]["lose_tyre"])
                ws.cell(row=row_idx, column=9, value=0)
                ws.cell(row=row_idx, column=10, value=row["aggregated"]["packing"])

        total_row = len(rows) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=totals["all_curing"]).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=totals["production_tyre"]).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=totals["repair"]).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=totals["second_grade"]).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=totals["third_grade"]).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=totals["lose_tyre"]).font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=totals["rfm"]).font = Font(bold=True)
        ws.cell(row=total_row, column=10, value=totals["packing"]).font = Font(bold=True)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="production_sheet_{range_label.replace(" ", "_")}.xlsx"'
        wb.save(response)
        return response

    context = {
        "rows": rows,
        "totals": totals,
        "date_str": date_str,
        "month_str": month_str,
        "range_label": range_label,
        "chart_labels": chart_labels,
        "chart_values": chart_values,
        "chart_repair": chart_repair,
        "chart_second": chart_second,
        "chart_third": chart_third,
        "chart_lose": chart_lose,
        "chart_curing": chart_curing,
        "prev_date": prev_date.strftime("%Y-%m-%d"),
        "next_date": next_date.strftime("%Y-%m-%d"),
        "page_obj": page_obj if date_str else None,
        "is_date_view": bool(date_str),
    }
    return render(request, "stock/production_sheet.html", context)